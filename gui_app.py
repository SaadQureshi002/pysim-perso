import json
import os
import re
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import messagebox, ttk

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from gsm_data_generator import DataGenerationScript, json_loader



REPO_ROOT = Path(__file__).resolve().parent
SETTINGS_FILE = REPO_ROOT / "settings.json"



NAVY = "#0B1F33"
NAVY_LIGHT = "#123653"
BLUE = "#1769E0"
BLUE_HOVER = "#0F58C7"

BACKGROUND = "#F4F7FA"
CARD = "#FFFFFF"

TEXT_PRIMARY = "#18212F"
TEXT_SECONDARY = "#6B7785"

BORDER = "#D9E0E7"
SUCCESS = "#138A58"
ERROR = "#C93C37"
WARNING = "#D58A00"

ENTRY_BG = "#F8FAFC"



last_excel_path = None
current_batch_id = None



def get_output_dir(config):
    output_dir = REPO_ROOT / config.PATHS.OUTPUT_FILES_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def get_history_path(config):
    return get_output_dir(config) / ".batch_history.json"


def load_history(config):
    history_path = get_history_path(config)

    if not history_path.exists():
        return []

    try:
        with open(history_path, "r", encoding="utf-8") as file:
            data = json.load(file)

        if isinstance(data, list):
            return data

    except Exception:
        pass

    return []


def save_history(config, history):
    history_path = get_history_path(config)

    with open(history_path, "w", encoding="utf-8") as file:
        json.dump(
            history,
            file,
            indent=4
        )


def generate_batch_id(config):
    history = load_history(config)

    today = datetime.now().strftime("%Y%m%d")

    today_batches = [
        item
        for item in history
        if str(item.get("batch_id", "")).startswith(f"BATCH-{today}-")
    ]

    number = len(today_batches) + 1

    return f"BATCH-{today}-{number:04d}"



def get_next_ranges(config):
    output_dir = get_output_dir(config)
    ledger_path = output_dir / ".issuance_ledger.json"

    default_iccid = str(config.DISP.iccid)
    default_imsi = str(config.DISP.imsi)

    if ledger_path.exists():
        try:
            with open(ledger_path, "r", encoding="utf-8") as file:
                ledger = json.load(file)

            if ledger:
                last_batch = ledger[-1]

                default_iccid = str(
                    int(last_batch["iccid_end"]) + 1
                )

                default_imsi = str(
                    int(last_batch["imsi_end"]) + 1
                )

        except Exception:
            pass

    return default_iccid, default_imsi


def calculate_end_range(start_value, quantity):
    try:
        return str(
            int(start_value) + int(quantity) - 1
        )

    except (ValueError, TypeError):
        return "-"


def update_range_preview(*args):
    start_iccid = iccid_var.get().strip()
    start_imsi = imsi_var.get().strip()
    size = size_var.get().strip()

    if (
        start_iccid.isdigit()
        and start_imsi.isdigit()
        and size.isdigit()
        and int(size) > 0
    ):
        end_iccid = calculate_end_range(
            start_iccid,
            size
        )

        end_imsi = calculate_end_range(
            start_imsi,
            size
        )

        end_iccid_value.config(
            text=end_iccid,
            fg=TEXT_PRIMARY
        )

        end_imsi_value.config(
            text=end_imsi,
            fg=TEXT_PRIMARY
        )

    else:
        end_iccid_value.config(
            text="-",
            fg=TEXT_SECONDARY
        )

        end_imsi_value.config(
            text="-",
            fg=TEXT_SECONDARY
        )



def set_status(message, status_type="normal"):
    status_var.set(message)

    if status_type == "success":
        status_dot.config(fg=SUCCESS)

    elif status_type == "error":
        status_dot.config(fg=ERROR)

    elif status_type == "working":
        status_dot.config(fg=BLUE)

    elif status_type == "warning":
        status_dot.config(fg=WARNING)

    else:
        status_dot.config(fg=TEXT_SECONDARY)


def sanitize_filename(name):
    name = name.strip()

    name = re.sub(
        r'[<>:"/\\|?*]',
        "_",
        name
    )

    return name or "SIM_BATCH"


def load_next_ranges():
    global current_batch_id

    try:
        config = json_loader(
            str(SETTINGS_FILE)
        )

        next_iccid, next_imsi = get_next_ranges(
            config
        )

        iccid_var.set(next_iccid)
        imsi_var.set(next_imsi)

        current_batch_id = generate_batch_id(
            config
        )

        batch_id_value.config(
            text=current_batch_id
        )

        update_range_preview()

    except Exception as exc:
        messagebox.showerror(
            "Configuration Error",
            str(exc)
        )


def clear_fields():
    try:
        config = json_loader(
            str(SETTINGS_FILE)
        )

        load_next_ranges()

        size_var.set(
            str(config.DISP.size)
        )

        filename_var.set(
            str(config.PATHS.FILE_NAME)
        )

        pin1_var.set(
            str(config.DISP.pin1)
        )

        puk1_var.set(
            str(config.DISP.puk1)
        )

        pin2_var.set(
            str(config.DISP.pin2)
        )

        puk2_var.set(
            str(config.DISP.puk2)
        )

        result_value.config(
            text="No batch generated in this session."
        )

        open_excel_button.config(
            state="disabled"
        )

        set_status(
            "System ready",
            "normal"
        )

        update_range_preview()

    except Exception as exc:
        messagebox.showerror(
            "Reset Error",
            str(exc)
        )


def open_output_folder():
    try:
        config = json_loader(
            str(SETTINGS_FILE)
        )

        output_dir = get_output_dir(
            config
        )

        os.startfile(output_dir)

    except Exception as exc:
        messagebox.showerror(
            "Folder Error",
            str(exc)
        )


def open_excel_file():
    global last_excel_path

    if (
        last_excel_path
        and Path(last_excel_path).exists()
    ):
        try:
            os.startfile(
                last_excel_path
            )

        except Exception as exc:
            messagebox.showerror(
                "Open Excel Error",
                str(exc)
            )

    else:
        messagebox.showwarning(
            "File Not Available",
            "Generate a batch first."
        )



def format_excel(
    excel_path,
    batch_id,
    start_iccid,
    end_iccid,
    start_imsi,
    end_imsi,
    batch_size
):
    workbook = load_workbook(excel_path)


    data_sheet = None

    if "ELECT" in workbook.sheetnames:
        data_sheet = workbook["ELECT"]
        data_sheet.title = "SIM_DATA"

    elif workbook.sheetnames:
        data_sheet = workbook[workbook.sheetnames[0]]
        data_sheet.title = "SIM_DATA"


    if "BATCH_INFO" in workbook.sheetnames:
        del workbook["BATCH_INFO"]

    batch_info = workbook.create_sheet("BATCH_INFO")
    batch_info.sheet_view.showGridLines = False

    batch_info["A1"] = "SIM BATCH INFORMATION"
    batch_info["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )
    batch_info["A1"].fill = PatternFill(
        "solid",
        fgColor="0B1F33"
    )
    batch_info["A1"].alignment = Alignment(
        vertical="center"
    )

    batch_info.merge_cells("A1:D2")

    details = [
        ("Batch ID", batch_id),
        (
            "Generated At",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ),
        ("Number of SIMs", batch_size),
        ("Starting ICCID", start_iccid),
        ("Ending ICCID", end_iccid),
        ("Starting IMSI", start_imsi),
        ("Ending IMSI", end_imsi),
        ("Export Format", "Microsoft Excel (.xlsx)"),
        ("Issuance Protection", "Enabled")
    ]

    row = 4

    for label, value in details:
        batch_info.cell(
            row=row,
            column=1,
            value=label
        )

        batch_info.cell(
            row=row,
            column=2,
            value=value
        )

        batch_info.cell(
            row=row,
            column=1
        ).font = Font(
            bold=True,
            color="18212F"
        )

        batch_info.cell(
            row=row,
            column=1
        ).fill = PatternFill(
            "solid",
            fgColor="EAF1F8"
        )

        batch_info.cell(
            row=row,
            column=1
        ).alignment = Alignment(
            vertical="center"
        )

        batch_info.cell(
            row=row,
            column=2
        ).alignment = Alignment(
            vertical="center"
        )

        row += 1

    batch_info.column_dimensions["A"].width = 24
    batch_info.column_dimensions["B"].width = 38


    for sheet in workbook.worksheets:

        if sheet.title == "BATCH_INFO":
            continue

        sheet.sheet_view.showGridLines = False

        sheet.freeze_panes = "A2"

        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.auto_filter.ref = sheet.dimensions


        for cell in sheet[1]:
            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = PatternFill(
                "solid",
                fgColor="0B1F33"
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        sheet.row_dimensions[1].height = 26


        header_map = {}

        for cell in sheet[1]:
            if cell.value is not None:
                header_map[
                    str(cell.value).strip().upper()
                ] = cell.column


        text_columns = [
            "ICCID",
            "IMSI",
            "PIN1",
            "PIN2",
            "PUK1",
            "PUK2",
            "ADM1",
            "ADM6",
            "ACC",
            "KI",
            "OPC",
            "EKI",
            "KIC1",
            "KID1",
            "KIK1",
            "KIC2",
            "KID2",
            "KIK2",
            "KIC3",
            "KID3",
            "KIK3"
        ]

        for column_name in text_columns:

            if column_name in header_map:
                column_number = header_map[column_name]

                for row_number in range(
                    2,
                    sheet.max_row + 1
                ):
                    cell = sheet.cell(
                        row=row_number,
                        column=column_number
                    )

                    if cell.value is not None:
                        cell.value = str(cell.value)
                        cell.number_format = "@"


        for row_number in range(
            2,
            sheet.max_row + 1
        ):
            for column_number in range(
                1,
                sheet.max_column + 1
            ):
                cell = sheet.cell(
                    row=row_number,
                    column=column_number
                )

                cell.alignment = Alignment(
                    vertical="center"
                )

                if row_number % 2 == 0:
                    cell.fill = PatternFill(
                        "solid",
                        fgColor="F5F8FB"
                    )


        for column_cells in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                value = (
                    ""
                    if cell.value is None
                    else str(cell.value)
                )

                max_length = max(
                    max_length,
                    len(value)
                )

            adjusted_width = min(
                max(max_length + 3, 12),
                40
            )

            sheet.column_dimensions[
                column_letter
            ].width = adjusted_width


        for row_number in range(
            2,
            sheet.max_row + 1
        ):
            sheet.row_dimensions[
                row_number
            ].height = 21


    if "SIM_DATA" in workbook.sheetnames:
        workbook.active = workbook.sheetnames.index(
            "SIM_DATA"
        )

    workbook.save(excel_path)



def add_batch_history(
    config,
    batch_id,
    excel_path,
    start_iccid,
    end_iccid,
    start_imsi,
    end_imsi,
    batch_size
):
    history = load_history(config)

    history.append(
        {
            "batch_id": batch_id,
            "timestamp": datetime.now().isoformat(
                timespec="seconds"
            ),
            "size": batch_size,
            "iccid_start": start_iccid,
            "iccid_end": end_iccid,
            "imsi_start": start_imsi,
            "imsi_end": end_imsi,
            "excel_file": str(
                excel_path
            )
        }
    )

    save_history(
        config,
        history
    )



def generate_batch():
    global last_excel_path
    global current_batch_id

    generate_button.config(
        state="disabled"
    )

    reset_button.config(
        state="disabled"
    )

    try:
        config = json_loader(
            str(SETTINGS_FILE)
        )

        start_iccid = iccid_var.get().strip()
        start_imsi = imsi_var.get().strip()

        batch_size = size_var.get().strip()
        file_name = filename_var.get().strip()

        pin1 = pin1_var.get().strip()
        puk1 = puk1_var.get().strip()

        pin2 = pin2_var.get().strip()
        puk2 = puk2_var.get().strip()


        if (
            not start_iccid
            or not start_imsi
            or not batch_size
            or not file_name
        ):
            messagebox.showerror(
                "Missing Information",
                "Please complete all required batch fields."
            )

            return

        try:
            batch_size_int = int(
                batch_size
            )

            if batch_size_int <= 0:
                raise ValueError

        except ValueError:

            messagebox.showerror(
                "Invalid Batch Size",
                "Number of SIMs must be a positive whole number."
            )

            return

        if not start_iccid.isdigit():
            messagebox.showerror(
                "Invalid ICCID",
                "Starting ICCID must contain digits only."
            )

            return

        if not start_imsi.isdigit():
            messagebox.showerror(
                "Invalid IMSI",
                "Starting IMSI must contain digits only."
            )

            return

        if pin1 and not pin1.isdigit():
            messagebox.showerror(
                "Invalid PIN1",
                "PIN1 must contain digits only."
            )

            return

        if pin2 and not pin2.isdigit():
            messagebox.showerror(
                "Invalid PIN2",
                "PIN2 must contain digits only."
            )

            return

        if puk1 and not puk1.isdigit():
            messagebox.showerror(
                "Invalid PUK1",
                "PUK1 must contain digits only."
            )

            return

        if puk2 and not puk2.isdigit():
            messagebox.showerror(
                "Invalid PUK2",
                "PUK2 must contain digits only."
            )

            return

        end_iccid = calculate_end_range(
            start_iccid,
            batch_size_int
        )

        end_imsi = calculate_end_range(
            start_imsi,
            batch_size_int
        )


        config.DISP.iccid = start_iccid
        config.DISP.imsi = start_imsi

        config.DISP.size = batch_size_int

        config.DISP.pin1 = pin1
        config.DISP.puk1 = puk1

        config.DISP.pin2 = pin2
        config.DISP.puk2 = puk2

        current_batch_id = generate_batch_id(
            config
        )

        batch_id_value.config(
            text=current_batch_id
        )

        safe_name = sanitize_filename(
            file_name
        )

        unique_name = (
            f"{safe_name}_"
            f"{current_batch_id}"
        )

        config.PATHS.FILE_NAME = unique_name

        set_status(
            "Generating secure SIM batch...",
            "working"
        )

        progress_bar.start(10)

        root.update_idletasks()


        script = DataGenerationScript(
            config
        )

        script.json_to_global_params()

        result_dfs, keys = (
            script.generate_all_data()
        )

        written = script.write_outputs(
            result_dfs
        )

        output_dir = get_output_dir(
            config
        )

        excel_path = (
            output_dir
            / f"{unique_name}.xlsx"
        )


        with pd.ExcelWriter(
            excel_path,
            engine="openpyxl"
        ) as writer:

            for (
                sheet_name,
                dataframe
            ) in result_dfs.items():

                dataframe.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False
                )


        format_excel(
            excel_path=excel_path,
            batch_id=current_batch_id,
            start_iccid=start_iccid,
            end_iccid=end_iccid,
            start_imsi=start_imsi,
            end_imsi=end_imsi,
            batch_size=batch_size_int
        )


        for path in written.values():

            txt_path = Path(path)

            if not txt_path.is_absolute():
                txt_path = (
                    REPO_ROOT
                    / txt_path
                )

            if txt_path.exists():
                txt_path.unlink()


        add_batch_history(
            config=config,
            batch_id=current_batch_id,
            excel_path=excel_path,
            start_iccid=start_iccid,
            end_iccid=end_iccid,
            start_imsi=start_imsi,
            end_imsi=end_imsi,
            batch_size=batch_size_int
        )

        last_excel_path = str(
            excel_path
        )

        progress_bar.stop()

        open_excel_button.config(
            state="normal"
        )

        set_status(
            "Batch generated successfully",
            "success"
        )

        result_value.config(
            text=(
                f"Batch ID: {current_batch_id}\n\n"
                f"Records: {batch_size_int}\n\n"
                f"ICCID Range:\n"
                f"{start_iccid}  →  {end_iccid}\n\n"
                f"IMSI Range:\n"
                f"{start_imsi}  →  {end_imsi}\n\n"
                f"Excel File:\n"
                f"{excel_path}"
            )
        )

        messagebox.showinfo(
            "Batch Generated",
            (
                "SIM personalization batch "
                "generated successfully.\n\n"
                f"Batch ID: {current_batch_id}\n"
                f"Records: {batch_size_int}\n\n"
                f"Excel File:\n{excel_path}"
            )
        )

        load_next_ranges()

    except Exception as exc:

        progress_bar.stop()

        set_status(
            "Batch generation failed",
            "error"
        )

        messagebox.showerror(
            "Generation Error",
            str(exc)
        )

    finally:

        generate_button.config(
            state="normal"
        )

        reset_button.config(
            state="normal"
        )



root = tk.Tk()

root.title(
    "SIM Provisioning Suite"
)

root.geometry(
    "1180x760"
)

root.minsize(
    1050,
    650
)

root.configure(
    bg=BACKGROUND
)



style = ttk.Style()

style.theme_use("clam")

style.configure(
    "TEntry",
    padding=8,
    fieldbackground=ENTRY_BG,
    bordercolor=BORDER,
    lightcolor=BORDER,
    darkcolor=BORDER,
    font=("Segoe UI", 10)
)

style.configure(
    "Primary.TButton",
    font=("Segoe UI Semibold", 10),
    padding=(22, 11),
    background=BLUE,
    foreground="white",
    borderwidth=0
)

style.map(
    "Primary.TButton",
    background=[
        ("active", BLUE_HOVER),
        ("disabled", "#9BBCEA")
    ]
)

style.configure(
    "Secondary.TButton",
    font=("Segoe UI", 10),
    padding=(18, 10),
    background="#FFFFFF",
    foreground=TEXT_PRIMARY,
    bordercolor=BORDER
)

style.map(
    "Secondary.TButton",
    background=[
        ("active", "#EEF3F8")
    ]
)

style.configure(
    "Horizontal.TProgressbar",
    troughcolor="#E6ECF2",
    background=BLUE,
    borderwidth=0,
    thickness=5
)



header = tk.Frame(
    root,
    bg=NAVY,
    height=88
)

header.pack(
    fill="x"
)

header.pack_propagate(False)

brand_frame = tk.Frame(
    header,
    bg=NAVY
)

brand_frame.pack(
    side="left",
    padx=38,
    pady=14
)

logo_box = tk.Label(
    brand_frame,
    text="SIM",
    bg=BLUE,
    fg="white",
    font=("Segoe UI Semibold", 11),
    width=5,
    height=2
)

logo_box.pack(
    side="left",
    padx=(0, 15)
)

brand_text = tk.Frame(
    brand_frame,
    bg=NAVY
)

brand_text.pack(
    side="left"
)

tk.Label(
    brand_text,
    text="SIM Provisioning Suite",
    bg=NAVY,
    fg="white",
    font=("Segoe UI Semibold", 19)
).pack(
    anchor="w"
)

tk.Label(
    brand_text,
    text="Secure Personalization & Batch Management",
    bg=NAVY,
    fg="#AFC3D5",
    font=("Segoe UI", 9)
).pack(
    anchor="w",
    pady=(2, 0)
)

version_label = tk.Label(
    header,
    text="ENTERPRISE  •  v2.0",
    bg=NAVY_LIGHT,
    fg="#D9E8F4",
    font=("Segoe UI Semibold", 8),
    padx=13,
    pady=7
)

version_label.pack(
    side="right",
    padx=38
)



main = tk.Frame(
    root,
    bg=BACKGROUND
)

main.pack(
    fill="both",
    expand=True,
    padx=34,
    pady=(10, 8)
)

top_title = tk.Frame(
    main,
    bg=BACKGROUND
)

top_title.pack(
    fill="x",
    pady=(0, 7)
)

tk.Label(
    top_title,
    text="Create Personalization Batch",
    bg=BACKGROUND,
    fg=TEXT_PRIMARY,
    font=("Segoe UI Semibold", 17)
).pack(
    anchor="w"
)

tk.Label(
    top_title,
    text=(
        "Configure subscriber identifiers, validate ranges "
        "and export a controlled Excel personalization batch."
    ),
    bg=BACKGROUND,
    fg=TEXT_SECONDARY,
    font=("Segoe UI", 9)
).pack(
    anchor="w",
    pady=(3, 0)
)



info_strip = tk.Frame(
    main,
    bg=CARD,
    highlightbackground=BORDER,
    highlightthickness=1
)

info_strip.pack(
    fill="x",
    pady=(0, 10)
)

info_inner = tk.Frame(
    info_strip,
    bg=CARD
)

info_inner.pack(
    fill="x",
    padx=20,
    pady=8
)

tk.Label(
    info_inner,
    text="Batch ID",
    bg=CARD,
    fg=TEXT_SECONDARY,
    font=("Segoe UI", 8)
).pack(
    side="left"
)

batch_id_value = tk.Label(
    info_inner,
    text="-",
    bg=CARD,
    fg=TEXT_PRIMARY,
    font=("Consolas", 10, "bold")
)

batch_id_value.pack(
    side="left",
    padx=(8, 30)
)

tk.Label(
    info_inner,
    text="Duplicate Protection",
    bg=CARD,
    fg=TEXT_SECONDARY,
    font=("Segoe UI", 8)
).pack(
    side="left"
)

tk.Label(
    info_inner,
    text="● ENABLED",
    bg=CARD,
    fg=SUCCESS,
    font=("Segoe UI Semibold", 8)
).pack(
    side="left",
    padx=(8, 0)
)



forms_container = tk.Frame(
    main,
    bg=BACKGROUND
)

forms_container.pack(
    fill="x"
)

forms_container.columnconfigure(
    0,
    weight=3
)

forms_container.columnconfigure(
    1,
    weight=2
)



batch_card = tk.Frame(
    forms_container,
    bg=CARD,
    highlightbackground=BORDER,
    highlightthickness=1
)

batch_card.grid(
    row=0,
    column=0,
    sticky="nsew",
    padx=(0, 10)
)

batch_inner = tk.Frame(
    batch_card,
    bg=CARD
)

batch_inner.pack(
    fill="both",
    expand=True,
    padx=24,
    pady=12
)

batch_inner.columnconfigure(
    1,
    weight=1
)

tk.Label(
    batch_inner,
    text="Batch Configuration",
    bg=CARD,
    fg=TEXT_PRIMARY,
    font=("Segoe UI Semibold", 12)
).grid(
    row=0,
    column=0,
    columnspan=2,
    sticky="w"
)

tk.Label(
    batch_inner,
    text="Subscriber numbering and export configuration.",
    bg=CARD,
    fg=TEXT_SECONDARY,
    font=("Segoe UI", 8)
).grid(
    row=1,
    column=0,
    columnspan=2,
    sticky="w",
    pady=(2, 14)
)

iccid_var = tk.StringVar()
imsi_var = tk.StringVar()
size_var = tk.StringVar(value="10")
filename_var = tk.StringVar(value="PAF_LTE_0012")

batch_fields = [
    (
        "Starting ICCID",
        "First ICCID allocated to this batch",
        iccid_var
    ),
    (
        "Starting IMSI",
        "First IMSI allocated to this batch",
        imsi_var
    ),
    (
        "Number of SIMs",
        "Number of records to generate",
        size_var
    ),
    (
        "Output Name",
        "Base name for the Excel workbook",
        filename_var
    )
]

for index, (
    label_text,
    helper_text,
    variable
) in enumerate(batch_fields):

    row = index + 2

    label_box = tk.Frame(
        batch_inner,
        bg=CARD
    )

    label_box.grid(
        row=row,
        column=0,
        sticky="w",
        padx=(0, 20),
        pady=4
    )

    tk.Label(
        label_box,
        text=label_text,
        bg=CARD,
        fg=TEXT_PRIMARY,
        font=("Segoe UI Semibold", 9)
    ).pack(
        anchor="w"
    )

    tk.Label(
        label_box,
        text=helper_text,
        bg=CARD,
        fg=TEXT_SECONDARY,
        font=("Segoe UI", 7)
    ).pack(
        anchor="w"
    )

    ttk.Entry(
        batch_inner,
        textvariable=variable
    ).grid(
        row=row,
        column=1,
        sticky="ew",
        pady=4
    )



security_card = tk.Frame(
    forms_container,
    bg=CARD,
    highlightbackground=BORDER,
    highlightthickness=1
)

security_card.grid(
    row=0,
    column=1,
    sticky="nsew",
    padx=(10, 0)
)

security_inner = tk.Frame(
    security_card,
    bg=CARD
)

security_inner.pack(
    fill="both",
    expand=True,
    padx=24,
    pady=12
)

tk.Label(
    security_inner,
    text="Security Parameters",
    bg=CARD,
    fg=TEXT_PRIMARY,
    font=("Segoe UI Semibold", 12)
).pack(
    anchor="w"
)

tk.Label(
    security_inner,
    text="PIN and PUK configuration for generated records.",
    bg=CARD,
    fg=TEXT_SECONDARY,
    font=("Segoe UI", 8)
).pack(
    anchor="w",
    pady=(2, 14)
)

security_grid = tk.Frame(
    security_inner,
    bg=CARD
)

security_grid.pack(
    fill="x"
)

security_grid.columnconfigure(
    0,
    weight=1
)

security_grid.columnconfigure(
    1,
    weight=1
)

pin1_var = tk.StringVar()
puk1_var = tk.StringVar()
pin2_var = tk.StringVar()
puk2_var = tk.StringVar()

security_fields = [
    ("PIN1", pin1_var, 0, 0),
    ("PUK1", puk1_var, 0, 1),
    ("PIN2", pin2_var, 1, 0),
    ("PUK2", puk2_var, 1, 1)
]

for (
    label_text,
    variable,
    row,
    column
) in security_fields:

    field_frame = tk.Frame(
        security_grid,
        bg=CARD
    )

    field_frame.grid(
        row=row,
        column=column,
        sticky="ew",
        padx=(
            (0, 8)
            if column == 0
            else (8, 0)
        ),
        pady=4
    )

    tk.Label(
        field_frame,
        text=label_text,
        bg=CARD,
        fg=TEXT_PRIMARY,
        font=("Segoe UI Semibold", 9)
    ).pack(
        anchor="w",
        pady=(0, 4)
    )

    ttk.Entry(
        field_frame,
        textvariable=variable
    ).pack(
        fill="x"
    )



preview_card = tk.Frame(
    main,
    bg=CARD,
    highlightbackground=BORDER,
    highlightthickness=1
)

preview_card.pack(
    fill="x",
    pady=(7, 0)
)

preview_inner = tk.Frame(
    preview_card,
    bg=CARD
)

preview_inner.pack(
    fill="x",
    padx=22,
    pady=9
)

tk.Label(
    preview_inner,
    text="Range Preview",
    bg=CARD,
    fg=TEXT_PRIMARY,
    font=("Segoe UI Semibold", 10)
).pack(
    side="left",
    padx=(0, 28)
)

tk.Label(
    preview_inner,
    text="Ending ICCID:",
    bg=CARD,
    fg=TEXT_SECONDARY,
    font=("Segoe UI", 8)
).pack(
    side="left"
)

end_iccid_value = tk.Label(
    preview_inner,
    text="-",
    bg=CARD,
    fg=TEXT_PRIMARY,
    font=("Consolas", 9, "bold")
)

end_iccid_value.pack(
    side="left",
    padx=(7, 30)
)

tk.Label(
    preview_inner,
    text="Ending IMSI:",
    bg=CARD,
    fg=TEXT_SECONDARY,
    font=("Segoe UI", 8)
).pack(
    side="left"
)

end_imsi_value = tk.Label(
    preview_inner,
    text="-",
    bg=CARD,
    fg=TEXT_PRIMARY,
    font=("Consolas", 9, "bold")
)

end_imsi_value.pack(
    side="left",
    padx=(7, 0)
)



action_bar = tk.Frame(
    main,
    bg=BACKGROUND
)

action_bar.pack(
    fill="x",
    pady=(7, 7)
)

generate_button = ttk.Button(
    action_bar,
    text="Generate Excel Batch",
    command=generate_batch,
    style="Primary.TButton"
)

generate_button.pack(
    side="left"
)

reset_button = ttk.Button(
    action_bar,
    text="Reset",
    command=clear_fields,
    style="Secondary.TButton"
)

reset_button.pack(
    side="left",
    padx=(10, 0)
)

folder_button = ttk.Button(
    action_bar,
    text="Open Output Folder",
    command=open_output_folder,
    style="Secondary.TButton"
)

folder_button.pack(
    side="left",
    padx=(10, 0)
)

open_excel_button = ttk.Button(
    action_bar,
    text="Open Excel",
    command=open_excel_file,
    style="Secondary.TButton",
    state="disabled"
)

open_excel_button.pack(
    side="left",
    padx=(10, 0)
)


result_card = tk.Frame(
    main,
    bg=CARD,
    highlightbackground=BORDER,
    highlightthickness=1
)

result_card.pack(
    fill="both",
    expand=True
)

result_card.configure(height=125)
result_card.pack_propagate(False)


result_inner = tk.Frame(
    result_card,
    bg=CARD
)

result_inner.pack(
    fill="both",
    expand=True,
    padx=22,
    pady=10
)

result_header = tk.Frame(
    result_inner,
    bg=CARD
)

result_header.pack(
    fill="x"
)

tk.Label(
    result_header,
    text="Batch Result",
    bg=CARD,
    fg=TEXT_PRIMARY,
    font=("Segoe UI Semibold", 11)
).pack(
    side="left"
)

status_frame = tk.Frame(
    result_header,
    bg=CARD
)

status_frame.pack(
    side="right"
)

status_dot = tk.Label(
    status_frame,
    text="●",
    fg=TEXT_SECONDARY,
    bg=CARD,
    font=("Segoe UI", 10)
)

status_dot.pack(
    side="left",
    padx=(0, 6)
)

status_var = tk.StringVar(
    value="System ready"
)

tk.Label(
    status_frame,
    textvariable=status_var,
    bg=CARD,
    fg=TEXT_SECONDARY,
    font=("Segoe UI", 8)
).pack(
    side="left"
)

progress_bar = ttk.Progressbar(
    result_inner,
    mode="indeterminate",
    style="Horizontal.TProgressbar"
)

progress_bar.pack(
    fill="x",
    pady=(8, 8)
)

result_value = tk.Label(
    result_inner,
    text="No batch generated in this session.",
    bg=CARD,
    fg=TEXT_SECONDARY,
    justify="left",
    anchor="nw",
    font=("Consolas", 9)
)

result_value.pack(
    fill="both",
    expand=True
)





footer = tk.Frame(
    root,
    bg="#E9EEF3",
    height=35
)

footer.pack(
    fill="x"
)

footer.pack_propagate(False)

tk.Label(
    footer,
    text=(
        "SIM Provisioning Suite  •  "
        "Excel Export  •  "
        "Issuance Protection Enabled"
    ),
    bg="#E9EEF3",
    fg=TEXT_SECONDARY,
    font=("Segoe UI", 8)
).pack(
    side="left",
    padx=35,
    pady=9
)

tk.Label(
    footer,
    text="Enterprise Build v2.0",
    bg="#E9EEF3",
    fg=TEXT_SECONDARY,
    font=("Segoe UI", 8)
).pack(
    side="right",
    padx=35,
    pady=9
)





iccid_var.trace_add(
    "write",
    update_range_preview
)

imsi_var.trace_add(
    "write",
    update_range_preview
)

size_var.trace_add(
    "write",
    update_range_preview
)




load_next_ranges()

config = json_loader(
    str(SETTINGS_FILE)
)

pin1_var.set(
    str(config.DISP.pin1)
)

puk1_var.set(
    str(config.DISP.puk1)
)

pin2_var.set(
    str(config.DISP.pin2)
)

puk2_var.set(
    str(config.DISP.puk2)
)

update_range_preview()




root.mainloop()